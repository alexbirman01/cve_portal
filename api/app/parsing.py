from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from typing import Any

import openpyxl
import pdfplumber


_CVE_RE = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)
_GHSA_RE = re.compile(r"GHSA-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}", re.IGNORECASE)
# Structured Vulnerability column: CVE or GitHub Security Advisory.
_VULN_ID_RE = re.compile(
    rf"^(?:{_CVE_RE.pattern}|{_GHSA_RE.pattern})$",
    re.IGNORECASE,
)
_VULN_ID_FIND_RE = re.compile(
    rf"(?:{_CVE_RE.pattern}|{_GHSA_RE.pattern})",
    re.IGNORECASE,
)

# Fallback regex for image:tag in free text (PDF / description).
# Image must contain at least one letter; allows optional space after colon.
_IMAGE_RE = re.compile(
    r"(?P<image>[a-zA-Z0-9._/-]*[a-zA-Z][a-zA-Z0-9._/-]*): ?(?P<tag>[a-zA-Z0-9._-]{1,128})"
)

# Column keyword aliases — order matters: first match wins.
_COL_ALIASES: dict[str, list[str]] = {
    "vuln":    ["vulnerability", "cve id", "cve"],
    "image":   ["image repository", "image repo", "image name", "container", "repository", "repo"],
    "tag":     ["tag", "version", "image tag"],
    "package": ["packages", "package name", "package"],
    "pkg_ver": ["package version", "pkg version", "version"],
    "fix":     ["fix status", "fix", "remediation", "fixed version"],
    "sev":     ["severity", "risk rating"],
    "score":   ["cvss score", "cvss", "score"],
}

_FIX_VER_RE = re.compile(r"fixed in\s+([^\s,;]+)", re.IGNORECASE)
# Version-like string: starts with a digit, contains at least one dot or dash
_PLAIN_VER_RE = re.compile(r"^\d[\w.\-+~^]*$")

_SEV_LABELS = ("CRITICAL", "HIGH", "MEDIUM", "MODERATE", "LOW", "NEGLIGIBLE", "UNKNOWN")
_SEV_RANK = {
    "CRITICAL": 5,
    "HIGH": 4,
    "MEDIUM": 3,
    "LOW": 2,
    "NEGLIGIBLE": 1,
    "UNKNOWN": 0,
}


# ─── dataclasses ─────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class ExtractedCve:
    cve_id: str
    source: str


@dataclass(frozen=True)
class ExtractedImage:
    image: str
    tag: str
    source: str


@dataclass(frozen=True)
class ExtractedPackage:
    cve_id: str
    package_name: str
    package_version: str | None   # installed/vulnerable version
    fixed_version: str | None     # from "Fix Status" column
    source: str


@dataclass(frozen=True)
class CveImageFact:
    """A directly-linked (CVE_ID, image, tag) fact extracted from a structured row."""
    cve_id: str
    image: str
    tag: str
    source: str
    severity: str | None = None  # customer-reported (Aqua), normalized uppercase
    score: str | None = None     # customer-reported CVSS score when present


@dataclass(frozen=True)
class ParsedAttachment:
    attachment_id: str
    filename: str
    mime_type: str | None
    status: str  # ok | unparsed | needs_ocr | error
    text_preview: str | None
    cves: list[ExtractedCve]
    images: list[ExtractedImage]          # legacy: free-text image extraction
    packages: list[ExtractedPackage]
    cve_image_facts: list[CveImageFact]   # structured: (CVE, image, tag) per row


# ─── helpers ─────────────────────────────────────────────────────────────────

def _find_col(header: tuple, alias_key: str) -> int | None:
    """Return the first column index whose header matches any alias for the given key.

    Prefers an exact header match (case-insensitive) before substring match so
    ``Severity`` wins over ``Custom Severity``.
    """
    keywords = _COL_ALIASES.get(alias_key, [])
    lowered = [(i, str(h).strip().lower() if h else "") for i, h in enumerate(header)]
    for kw in keywords:
        for i, h in lowered:
            if h == kw:
                return i
    for kw in keywords:
        for i, h in lowered:
            if h and kw in h:
                return i
    return None


def normalize_customer_severity(raw: Any) -> str | None:
    """Map Aqua/customer severity strings to CRITICAL/HIGH/MEDIUM/LOW/NEGLIGIBLE."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    u = s.upper()
    for label in _SEV_LABELS:
        if label in u:
            return "MEDIUM" if label == "MODERATE" else label
    return None


def normalize_customer_score(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("none", "n/a", "-", "null"):
        return None
    try:
        float(s)
    except ValueError:
        return None
    return s


def severity_rank(sev: str | None) -> int:
    if not sev:
        return -1
    return _SEV_RANK.get(sev.upper().strip(), -1)


def _aqua_json_vuln_severity(vuln: dict[str, Any]) -> str | None:
    for key in ("aqua_severity", "nvd_severity", "vendor_severity", "severity"):
        sev = normalize_customer_severity(vuln.get(key))
        if sev:
            return sev
    return None


def _aqua_json_vuln_score(vuln: dict[str, Any]) -> str | None:
    for key in (
        "nvd_score_v3",
        "nvd_score_v2",
        "aqua_score",
        "vendor_score",
        "score",
        "nvd_score",
    ):
        score = normalize_customer_score(vuln.get(key))
        if score:
            return score
    return None


def _parse_fix_version(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    # Pattern 1: "Fixed in 3.8.13"
    m = _FIX_VER_RE.search(s)
    if m:
        return m.group(1)
    # Pattern 2: bare version string like "3.8.13" or "1.2.3-4"
    if _PLAIN_VER_RE.match(s):
        return s
    return None


# ─── public extraction helpers ───────────────────────────────────────────────

def normalize_vuln_id(raw: str) -> str | None:
    """Return normalized CVE/GHSA id, or None if not a supported vuln id."""
    s = (raw or "").strip()
    if not s or not _VULN_ID_RE.match(s):
        return None
    return s.upper()


def is_ghsa_id(vuln_id: str) -> bool:
    return bool(_GHSA_RE.fullmatch((vuln_id or "").strip()))


def is_cve_id(vuln_id: str) -> bool:
    return bool(_CVE_RE.fullmatch((vuln_id or "").strip()))


def extract_cves(text: str, source: str) -> list[ExtractedCve]:
    """Extract CVE and GHSA identifiers from free text (normalized uppercase)."""
    return [
        ExtractedCve(cve_id=m.group(0).upper(), source=source)
        for m in _VULN_ID_FIND_RE.finditer(text or "")
    ]


def cve_ids_from_attachment_facts(parsed_attachments: list[dict]) -> list[str]:
    """Return sorted unique vuln IDs from structured Excel/Aqua JSON facts only.

    Description and PDF free-text CVE mentions must not create findings rows.
    ``parsed_attachments`` entries are dicts with a ``cve_image_facts`` list
    (as produced by the worker after ``parse_attachment_bytes``).
    """
    return sorted(
        {
            f["cve_id"]
            for p in parsed_attachments
            for f in (p.get("cve_image_facts") or [])
            if isinstance(f, dict) and f.get("cve_id")
        }
    )


def extract_images(text: str, source: str) -> list[ExtractedImage]:
    out: list[ExtractedImage] = []
    for m in _IMAGE_RE.finditer(text or ""):
        out.append(ExtractedImage(image=m.group("image"), tag=m.group("tag"), source=source))
    return out


def list_excel_sheets(data: bytes) -> list[dict[str, Any]]:
    """Return sheet metadata for multi-sheet selection UI.

    Each entry: ``{"name": str, "row_count": int}`` (data rows, excluding header).
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    try:
        for sheet in wb.worksheets:
            # Cheap row count: iterate once; read_only worksheets support this.
            n = 0
            for i, _row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue  # header
                n += 1
            out.append({"name": sheet.title, "row_count": n})
    finally:
        wb.close()
    return out


# ─── PDF parser ──────────────────────────────────────────────────────────────

def parse_pdf_bytes(data: bytes, source: str, attachment_id: str, filename: str, mime_type: str | None) -> ParsedAttachment:
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            texts: list[str] = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                if t:
                    texts.append(t)
            full_text = "\n\n".join(texts).strip()
    except Exception as e:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="error", text_preview=str(e)[:1000],
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    if not full_text:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="needs_ocr", text_preview=None,
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    cves   = extract_cves(full_text, source)
    images = extract_images(full_text, source)
    return ParsedAttachment(
        attachment_id=attachment_id, filename=filename, mime_type=mime_type,
        status="ok", text_preview=full_text[:2000],
        cves=cves, images=images, packages=[], cve_image_facts=[],
    )


# ─── Excel parser ─────────────────────────────────────────────────────────────

def parse_excel_bytes(
    data: bytes,
    source: str,
    attachment_id: str,
    filename: str,
    mime_type: str | None,
    sheet_names: set[str] | None = None,
) -> ParsedAttachment:
    try:
        # When filtering sheets, avoid read_only: skipping unread worksheets can
        # desync openpyxl's XML stream and hang the worker indefinitely.
        use_read_only = sheet_names is None
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=use_read_only, data_only=True)

        # Collect structured facts; use a set to dedup (CVE, image, tag) across sheets.
        seen_facts: set[tuple[str, str, str]] = set()
        cve_image_facts: list[CveImageFact] = []
        packages: list[ExtractedPackage] = []
        seen_pkgs: set[tuple[str, str]] = set()  # (cve_id, package_name)
        texts: list[str] = []

        selected = {n.strip() for n in sheet_names} if sheet_names is not None else None

        for sheet in wb.worksheets:
            if selected is not None and sheet.title not in selected:
                continue

            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            header = rows[0]

            # Detect column positions using alias table.
            vuln_col = _find_col(header, "vuln")
            img_col  = _find_col(header, "image")
            tag_col  = _find_col(header, "tag")
            pkg_col  = _find_col(header, "package")
            ver_col  = _find_col(header, "pkg_ver")
            fix_col  = _find_col(header, "fix")
            sev_col  = _find_col(header, "sev")
            score_col = _find_col(header, "score")

            for row in rows:
                # Always collect raw text for fallback CVE regex extraction.
                row_vals = [str(v) for v in row if v is not None]
                if row_vals:
                    texts.append(" | ".join(row_vals))

                if row is rows[0]:
                    continue  # skip header row for structured extraction

                # Need at minimum a CVE/GHSA column to do structured extraction.
                if vuln_col is None:
                    continue

                cve_raw = str(row[vuln_col]).strip() if row[vuln_col] else ""
                cve_id = normalize_vuln_id(cve_raw)
                if not cve_id:
                    continue

                cust_sev = (
                    normalize_customer_severity(row[sev_col])
                    if sev_col is not None and row[sev_col] is not None
                    else None
                )
                cust_score = (
                    normalize_customer_score(row[score_col])
                    if score_col is not None and row[score_col] is not None
                    else None
                )

                # ── (CVE/GHSA, image, tag) fact ──
                if img_col is not None and row[img_col]:
                    img_name = str(row[img_col]).strip()
                    img_tag  = str(row[tag_col]).strip() if tag_col is not None and row[tag_col] else ""
                    key = (cve_id, img_name, img_tag)
                    if img_name and key not in seen_facts:
                        seen_facts.add(key)
                        cve_image_facts.append(CveImageFact(
                            cve_id=cve_id,
                            image=img_name,
                            tag=img_tag,
                            source=source,
                            severity=cust_sev,
                            score=cust_score,
                        ))

                # ── package fact ──
                if pkg_col is not None and row[pkg_col]:
                    pkg_name = str(row[pkg_col]).strip()
                    pkg_key  = (cve_id, pkg_name)
                    if pkg_key not in seen_pkgs:
                        seen_pkgs.add(pkg_key)
                        pkg_ver = str(row[ver_col]).strip() if ver_col is not None and row[ver_col] else None
                        fix_raw = str(row[fix_col]).strip() if fix_col is not None and row[fix_col] else None
                        packages.append(ExtractedPackage(
                            cve_id=cve_id,
                            package_name=pkg_name,
                            package_version=pkg_ver or None,
                            fixed_version=_parse_fix_version(fix_raw),
                            source=source,
                        ))

        full_text = "\n".join(texts)

        # Always extract CVEs/GHSAs from raw text (catches any format not in known columns).
        cves = extract_cves(full_text, source)

        # Images: use free-text regex only as fallback when structured extraction found nothing.
        images = [] if cve_image_facts else extract_images(full_text, source)

    except Exception as e:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="error", text_preview=str(e)[:1000],
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    return ParsedAttachment(
        attachment_id=attachment_id, filename=filename, mime_type=mime_type,
        status="ok", text_preview=full_text[:2000],
        cves=cves, images=images, packages=packages, cve_image_facts=cve_image_facts,
    )


# ─── Aqua Security JSON parser ────────────────────────────────────────────────

# ECR transactional tags: {version}_{Service}_{suffix}
# suffix is usually DDMonYYYY (e.g. 15Jun2026) but can be non-calendar
# (e.g. Junhotfix2026). Service may contain underscores (pip_operator).
_AQUA_TAG_SERVICE_RE = re.compile(r"^([\d.]+)_(.+)_(.+)$")


def _aqua_image_basename(image_name: str) -> tuple[str, str]:
    """Return (service_token, tag) extracted from a full Aqua image_name field.

    image_name format:
      registry/repo:version_ServiceName_suffix  (ECR transactional repo)
      registry/repo:tag                         (standard naming)

    The service token is the segment between the version and trailing suffix when
    the tag follows the Aqua transactional pattern; otherwise it is the repo path.
    Underscores in the service token are normalized to hyphens for catalog matching
    (e.g. pip_operator → pip-operator).
    """
    s = image_name.strip()
    # Strip registry prefix (first component containing "." or port ":")
    parts = s.split("/", 1)
    if len(parts) == 2 and ("." in parts[0] or (parts[0].count(":") == 1 and parts[0].split(":")[1].isdigit())):
        s = parts[1]
    # Split tag
    if ":" in s:
        repo, tag = s.rsplit(":", 1)
    else:
        repo, tag = s, ""
    # Try to extract a service identifier from {version}_{Service}_{suffix}
    m = _AQUA_TAG_SERVICE_RE.match(tag)
    if m:
        service = m.group(2).lower().replace("_", "-")
    else:
        service = repo.strip().lower()
    return service, tag.strip()


def parse_aqua_json_bytes(
    data: bytes,
    source: str,
    attachment_id: str,
    filename: str,
    mime_type: str | None,
    alias_map: dict[str, str] | None = None,
) -> ParsedAttachment:
    """Parse an Aqua Security scan report JSON attachment.

    Supports the format exported by Aqua Cloud (array of scan results with
    ``image_name`` and ``results.resources[].vulnerabilities[]``).
    ``alias_map`` resolves vendor-specific service tokens to canonical image basenames.
    """
    try:
        raw = json.loads(data.decode("utf-8", errors="replace"))
    except Exception as exc:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="error", text_preview=str(exc)[:500],
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    # Validate Aqua shape: list of scan entries with image_name + results.resources
    if not isinstance(raw, list) or not raw:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="unparsed", text_preview=None,
            cves=[], images=[], packages=[], cve_image_facts=[],
        )
    first = raw[0] if isinstance(raw[0], dict) else {}
    if "image_name" not in first or "results" not in first:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="unparsed", text_preview=None,
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    seen_facts: set[tuple[str, str, str]] = set()
    cve_image_facts: list[CveImageFact] = []
    packages: list[ExtractedPackage] = []
    seen_pkgs: set[tuple[str, str]] = set()
    all_cve_ids: list[str] = []

    for entry in raw:
        if not isinstance(entry, dict):
            continue
        image_name = str(entry.get("image_name") or "")
        service_token, tag = _aqua_image_basename(image_name)
        # Resolve alias → canonical name; fall back to the token itself
        canonical = (alias_map or {}).get(service_token, service_token)

        resources = (entry.get("results") or {}).get("resources") or []
        for res in resources:
            if not isinstance(res, dict):
                continue
            pkg = res.get("resource") or {}
            pkg_name = str(pkg.get("name") or "").strip()
            pkg_version = str(pkg.get("version") or "").strip() or None

            for vuln in (res.get("vulnerabilities") or []):
                if not isinstance(vuln, dict):
                    continue
                cve_raw = str(vuln.get("name") or "").strip()
                cve_id = normalize_vuln_id(cve_raw)
                if not cve_id:
                    continue
                all_cve_ids.append(cve_id)
                cust_sev = _aqua_json_vuln_severity(vuln)
                cust_score = _aqua_json_vuln_score(vuln)

                key = (cve_id, canonical, tag)
                if key not in seen_facts:
                    seen_facts.add(key)
                    cve_image_facts.append(CveImageFact(
                        cve_id=cve_id,
                        image=canonical,
                        tag=tag,
                        source=source,
                        severity=cust_sev,
                        score=cust_score,
                    ))

                if pkg_name:
                    pkg_key = (cve_id, pkg_name)
                    if pkg_key not in seen_pkgs:
                        seen_pkgs.add(pkg_key)
                        packages.append(ExtractedPackage(
                            cve_id=cve_id,
                            package_name=pkg_name,
                            package_version=pkg_version,
                            fixed_version=None,
                            source=source,
                        ))

    cves = [ExtractedCve(cve_id=c, source=source) for c in dict.fromkeys(all_cve_ids)]
    return ParsedAttachment(
        attachment_id=attachment_id, filename=filename, mime_type=mime_type,
        status="ok", text_preview=None,
        cves=cves, images=[], packages=packages, cve_image_facts=cve_image_facts,
    )


# ─── Aqua Security HTML report parser ─────────────────────────────────────────

_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")
_AQUA_HTML_H1_RE = re.compile(
    r"<h1[^>]*>\s*Scan Report:\s*(.*?)</h1>",
    re.IGNORECASE | re.DOTALL,
)
_AQUA_HTML_TITLE_RE = re.compile(
    r"<title[^>]*>(.*?)</title>",
    re.IGNORECASE | re.DOTALL,
)


def _strip_html_text(raw: str) -> str:
    return _WS_RE.sub(" ", _TAG_RE.sub(" ", raw or "")).strip()


def _html_tables(html: str) -> list[tuple[list[str], list[list[str]]]]:
    """Return ``[(headers, data_rows), ...]`` for each ``<table>`` in ``html``."""
    out: list[tuple[list[str], list[list[str]]]] = []
    for table_html in re.findall(r"<table[^>]*>(.*?)</table>", html, re.IGNORECASE | re.DOTALL):
        headers = [
            _strip_html_text(h)
            for h in re.findall(r"<th[^>]*>(.*?)</th>", table_html, re.IGNORECASE | re.DOTALL)
        ]
        data_rows: list[list[str]] = []
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", table_html, re.IGNORECASE | re.DOTALL):
            cells = [
                _strip_html_text(c)
                for c in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.IGNORECASE | re.DOTALL)
            ]
            if cells:
                data_rows.append(cells)
        if headers or data_rows:
            out.append((headers, data_rows))
    return out


def _aqua_html_image_ref(html: str) -> str | None:
    """Extract full ``registry/.../image:tag`` from Aqua HTML H1 or title."""
    m = _AQUA_HTML_H1_RE.search(html)
    if m:
        ref = _strip_html_text(m.group(1))
        if ref:
            return ref
    m = _AQUA_HTML_TITLE_RE.search(html)
    if not m:
        return None
    title = _strip_html_text(m.group(1))
    if "|" in title:
        title = title.split("|", 1)[0].strip()
    return title or None


def parse_aqua_html_bytes(
    data: bytes,
    source: str,
    attachment_id: str,
    filename: str,
    mime_type: str | None,
    alias_map: dict[str, str] | None = None,
) -> ParsedAttachment:
    """Parse an Aqua Security HTML scan report (``AquaReport-*.html``).

    Image/tag come from ``Scan Report: …``; findings prefer the consolidated
    table with both ``Name`` and ``Resource`` columns (severity-bucket tables
    are fallbacks when that table is absent).
    """
    try:
        html = data.decode("utf-8", errors="replace")
    except Exception as exc:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="error", text_preview=str(exc)[:500],
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    image_ref = _aqua_html_image_ref(html)
    if not image_ref:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="unparsed", text_preview=html[:500] or None,
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    service_token, tag = _aqua_image_basename(image_ref)
    canonical = (alias_map or {}).get(service_token, service_token)

    consolidated: list[dict[str, str | None]] = []
    name_only: list[dict[str, str | None]] = []
    for headers, rows in _html_tables(html):
        hl = [h.lower() for h in headers]
        if "name" not in hl:
            continue
        name_i = hl.index("name")
        res_i = hl.index("resource") if "resource" in hl else None
        fix_i = hl.index("fix version") if "fix version" in hl else None
        sev_i = hl.index("severity") if "severity" in hl else None
        score_i = hl.index("score") if "score" in hl else None
        for cells in rows:
            if name_i >= len(cells):
                continue
            cve_id = normalize_vuln_id(cells[name_i])
            if not cve_id:
                continue
            resource = None
            if res_i is not None and res_i < len(cells):
                resource = cells[res_i].strip() or None
            fixed = None
            if fix_i is not None and fix_i < len(cells):
                raw_fix = cells[fix_i].strip()
                if raw_fix and raw_fix.lower() != "none":
                    fixed = raw_fix
            cust_sev = (
                normalize_customer_severity(cells[sev_i])
                if sev_i is not None and sev_i < len(cells)
                else None
            )
            cust_score = (
                normalize_customer_score(cells[score_i])
                if score_i is not None and score_i < len(cells)
                else None
            )
            entry = {
                "cve_id": cve_id,
                "resource": resource,
                "fixed_version": fixed,
                "severity": cust_sev,
                "score": cust_score,
            }
            if res_i is not None:
                consolidated.append(entry)
            else:
                name_only.append(entry)

    rows_src = consolidated if consolidated else name_only
    if not rows_src:
        return ParsedAttachment(
            attachment_id=attachment_id, filename=filename, mime_type=mime_type,
            status="unparsed", text_preview=html[:500] or None,
            cves=[], images=[], packages=[], cve_image_facts=[],
        )

    seen_facts: set[tuple[str, str, str]] = set()
    cve_image_facts: list[CveImageFact] = []
    packages: list[ExtractedPackage] = []
    seen_pkgs: set[tuple[str, str]] = set()
    all_cve_ids: list[str] = []

    for row in rows_src:
        cve_id = str(row["cve_id"])
        all_cve_ids.append(cve_id)
        key = (cve_id, canonical, tag)
        if key not in seen_facts:
            seen_facts.add(key)
            cve_image_facts.append(
                CveImageFact(
                    cve_id=cve_id,
                    image=canonical,
                    tag=tag,
                    source=source,
                    severity=row.get("severity"),
                    score=row.get("score"),
                )
            )
        pkg_name = row.get("resource")
        if pkg_name:
            pkg_key = (cve_id, pkg_name)
            if pkg_key not in seen_pkgs:
                seen_pkgs.add(pkg_key)
                packages.append(
                    ExtractedPackage(
                        cve_id=cve_id,
                        package_name=pkg_name,
                        package_version=None,
                        fixed_version=row.get("fixed_version"),
                        source=source,
                    )
                )

    cves = [ExtractedCve(cve_id=c, source=source) for c in dict.fromkeys(all_cve_ids)]
    preview = f"Aqua HTML: {canonical}:{tag} — {len(cve_image_facts)} finding(s)"
    return ParsedAttachment(
        attachment_id=attachment_id, filename=filename, mime_type=mime_type,
        status="ok", text_preview=preview,
        cves=cves, images=[], packages=packages, cve_image_facts=cve_image_facts,
    )


# ─── dispatcher ──────────────────────────────────────────────────────────────

def parse_attachment_bytes(
    *,
    attachment_id: str,
    filename: str,
    mime_type: str | None,
    data: bytes,
    alias_map: dict[str, str] | None = None,
    sheet_names: set[str] | None = None,
) -> ParsedAttachment:
    source = f"attachment:{attachment_id}:{filename}"
    lower  = filename.lower()

    if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")) or mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return parse_excel_bytes(
            data, source, attachment_id, filename, mime_type, sheet_names=sheet_names,
        )

    if lower.endswith(".pdf") or mime_type == "application/pdf":
        return parse_pdf_bytes(data, source, attachment_id, filename, mime_type)

    if lower.endswith(".json"):
        return parse_aqua_json_bytes(data, source, attachment_id, filename, mime_type, alias_map=alias_map)

    if lower.endswith((".html", ".htm")) or (mime_type or "").split(";")[0].strip().lower() in (
        "text/html",
        "application/xhtml+xml",
    ):
        return parse_aqua_html_bytes(
            data, source, attachment_id, filename, mime_type, alias_map=alias_map,
        )

    return ParsedAttachment(
        attachment_id=attachment_id, filename=filename, mime_type=mime_type,
        status="unparsed", text_preview=None,
        cves=[], images=[], packages=[], cve_image_facts=[],
    )


# ─── ADF / Jira description helpers ──────────────────────────────────────────

def _adf_node_to_text(node: Any) -> str:
    """Recursively convert an Atlassian Document Format node to plain text."""
    if not isinstance(node, dict):
        return ""
    ntype = node.get("type", "")
    if ntype == "text":
        return node.get("text", "")
    if ntype == "hardBreak":
        return "\n"
    children = node.get("content") or []
    if ntype == "tableRow":
        cells = [_adf_node_to_text(c).rstrip("\n") for c in children]
        return " | ".join(cells) + "\n"
    if ntype in ("tableHeader", "tableCell"):
        return "".join(_adf_node_to_text(c) for c in children)
    if ntype == "table":
        return "".join(_adf_node_to_text(r) for r in children)
    inner = "".join(_adf_node_to_text(c) for c in children)
    if ntype in ("paragraph", "heading", "listItem"):
        return inner + "\n"
    if ntype in ("bulletList", "orderedList"):
        return inner
    return inner


def normalize_description(description_raw: Any) -> str:
    """Return a plain-text representation suitable for CVE extraction and display."""
    if description_raw is None:
        return ""
    if isinstance(description_raw, dict) and description_raw.get("type") == "doc":
        return _adf_node_to_text(description_raw).strip()
    return str(description_raw)
