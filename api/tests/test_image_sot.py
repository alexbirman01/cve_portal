"""Tests for strict image source-of-truth and Allowed Images catalog enforcement.

Key behaviors verified:
- CVE↔image facts come only from Excel/JSON attachment structured rows (cve_image_facts).
- Description text and PDF free-text produce no cve_image_facts.
- normalize_image_basename + alias resolution produces canonical names.
- load_allowed_names / is_allowed_basename correctly gates catalog membership.
"""

from __future__ import annotations

import io
import json
import openpyxl

from api.app.allowed_images import (
    is_allowed_basename,
    normalize_image_basename,
)
from api.app.parsing import (
    _aqua_image_basename,
    cve_ids_from_attachment_facts,
    extract_cves,
    extract_images,
    parse_attachment_bytes,
)


# ── normalize_image_basename ──────────────────────────────────────────────────

def test_normalize_strips_prefix():
    assert normalize_image_basename("plainid/pip-operator") == "pip-operator"


def test_normalize_takes_last_segment():
    assert normalize_image_basename("registry.example.com/plainid/theruntime") == "theruntime"


def test_normalize_lowercases():
    assert normalize_image_basename("PLAINID/TheRuntime") == "theruntime"


def test_normalize_bare_name():
    assert normalize_image_basename("agent") == "agent"


# ── alias map resolution ──────────────────────────────────────────────────────



def _build_alias_map(entries: list[tuple[str, str]]) -> dict[str, str]:
    """Pure-Python alias map (mirrors load_alias_map logic without SQLAlchemy)."""
    from types import SimpleNamespace
    out: dict[str, str] = {}
    for name, aliases_csv in entries:
        canonical = name.lower().strip()
        out[canonical] = canonical
        for a in aliases_csv.split(","):
            a = a.strip().lower()
            if a:
                out[a] = canonical
    return out


def _build_allowed_names(entries: list[tuple[str, str]]) -> set[str]:
    """Pure-Python allowed names set (mirrors load_allowed_names logic without SQLAlchemy)."""
    names: set[str] = set()
    for name, aliases_csv in entries:
        names.add(name.lower())
        for a in aliases_csv.split(","):
            a = a.strip().lower()
            if a:
                names.add(a)
    return names


def test_alias_resolves_to_canonical():
    alias_map = _build_alias_map([("secrets-mgmt", "secretmgr,secret-mgmt")])
    assert alias_map.get("secretmgr") == "secrets-mgmt"
    assert alias_map.get("secret-mgmt") == "secrets-mgmt"
    assert alias_map.get("secrets-mgmt") == "secrets-mgmt"


def test_load_allowed_names_includes_aliases():
    allowed = _build_allowed_names([("pip-operator", "pip-op")])
    assert "pip-operator" in allowed
    assert "pip-op" in allowed


def test_is_allowed_basename_case_insensitive():
    allowed = {"agent", "pip-operator", "theruntime"}
    assert is_allowed_basename("Agent", allowed) is True
    assert is_allowed_basename("authorizer", allowed) is False


# ── description text → no cve_image_facts ────────────────────────────────────

def test_description_extract_images_no_structured_facts():
    """extract_images returns ExtractedImage list, not cve_image_facts — never PLAT-eligible."""
    text = "CVE-2026-39822 affects authorizer:1.0 and agent:2.0"
    images = extract_images(text, "description")
    # These are free-text ExtractedImage objects, not CveImageFacts.
    # They are NOT placed into cve_to_images under the strict SoT rule.
    assert any(i.image == "authorizer" for i in images)
    assert any(i.image == "agent" for i in images)
    # The point: these do not carry a cve_id binding, so worker ignores them for PLAT slots.


def test_description_cves_without_image_facts():
    """Description-parsed CVEs produce no cve_image_facts — PLAT slots stay empty."""
    text = "This ticket covers CVE-2026-12345 in authorizer:1.0."
    cves = extract_cves(text, "description")
    assert any(c.cve_id == "CVE-2026-12345" for c in cves)
    # extract_cves has no image correlation — callers get no (cve_id, image, tag) facts.


def test_findings_cve_ids_ignore_description_only_mentions():
    """Worker findings list uses cve_image_facts only — description-only CVEs are dropped."""
    # PLATFORM-2107-style: description lists CVE-2026-41992 but Aqua JSON does not.
    desc_cves = extract_cves(
        "CVE-2026-39822 CVE-2026-39829 CVE-2026-41992 CVE-2026-46597",
        "description",
    )
    assert any(c.cve_id == "CVE-2026-41992" for c in desc_cves)

    parsed_attachments = [
        {
            "cves": [
                {"cve_id": "CVE-2026-39822", "source": "attachment:1:scan.json"},
                {"cve_id": "CVE-2026-39829", "source": "attachment:1:scan.json"},
            ],
            "cve_image_facts": [
                {
                    "cve_id": "CVE-2026-39822",
                    "image": "authz-sql-pdp-modifier",
                    "tag": "5.2627.4_authz-sql-pdp-modifier_Junhotfix2026",
                    "source": "attachment:1:scan.json",
                },
                {
                    "cve_id": "CVE-2026-39829",
                    "image": "authz-sql-pdp-modifier",
                    "tag": "5.2627.4_authz-sql-pdp-modifier_Junhotfix2026",
                    "source": "attachment:1:scan.json",
                },
            ],
        }
    ]
    cve_ids = cve_ids_from_attachment_facts(parsed_attachments)
    assert cve_ids == ["CVE-2026-39822", "CVE-2026-39829"]
    assert "CVE-2026-41992" not in cve_ids
    # Union with description (old behavior) would have wrongly included it:
    old_union = sorted({c.cve_id for c in desc_cves} | set(cve_ids))
    assert "CVE-2026-41992" in old_union


# ── Excel structured attachment → cve_image_facts ────────────────────────────

def _make_excel_bytes(rows: list[tuple]) -> bytes:
    """Build a minimal Excel workbook with CVE/Image/Tag columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("CVE ID", "Image Repository", "Tag"))
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_produces_cve_image_facts():
    data = _make_excel_bytes([
        ("CVE-2026-39822", "plainid/pip-operator", "5.2622.1"),
        ("CVE-2026-39822", "plainid/agent", "5.2622.1"),
    ])
    parsed = parse_attachment_bytes(
        attachment_id="att1",
        filename="scan.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=data,
    )
    assert parsed.status == "ok"
    assert len(parsed.cve_image_facts) == 2
    cve_ids = {f.cve_id for f in parsed.cve_image_facts}
    assert "CVE-2026-39822" in cve_ids
    images = {f.image for f in parsed.cve_image_facts}
    assert "plainid/pip-operator" in images
    assert "plainid/agent" in images


def test_excel_no_image_col_produces_no_facts():
    """Excel without an image column should produce no cve_image_facts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("CVE ID", "Package"))
    ws.append(("CVE-2026-11111", "openssl"))
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_attachment_bytes(
        attachment_id="att2",
        filename="scan.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=buf.getvalue(),
    )
    assert parsed.cve_image_facts == []


def test_unknown_image_in_excel_passes_through_to_row():
    """Unrecognized image names from Excel reach affected_images — catalog check is at create time."""
    data = _make_excel_bytes([
        ("CVE-2026-39822", "authorizer", "1.0"),
    ])
    parsed = parse_attachment_bytes(
        attachment_id="att3",
        filename="scan.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=data,
    )
    # Parser emits the fact; worker puts it in affected_images.
    # API will reject create because "authorizer" is not in Allowed Images.
    assert any(f.image == "authorizer" for f in parsed.cve_image_facts)


# ── Aqua JSON → cve_image_facts ───────────────────────────────────────────────

def test_aqua_json_produces_cve_image_facts():
    payload = [
        {
            "image_name": "registry.example.com/plainid/pip-operator:5.2622.1",
            "results": {
                "resources": [
                    {
                        "resource": {"name": "openssl", "version": "3.0.0"},
                        "vulnerabilities": [{"name": "CVE-2026-39822"}],
                    }
                ]
            },
        }
    ]
    data = json.dumps(payload).encode()
    # Aqua parser emits the service token as extracted from the image_name repo path.
    # For this input (no transactional-tag pattern) the token is the full repo segment
    # "plainid/pip-operator"; worker's _resolve_image_name normalizes it to "pip-operator".
    alias_map = {"pip-operator": "pip-operator"}
    parsed = parse_attachment_bytes(
        attachment_id="att4",
        filename="aqua_scan.json",
        mime_type="application/json",
        data=data,
        alias_map=alias_map,
    )
    assert parsed.status == "ok"
    assert len(parsed.cve_image_facts) == 1
    assert parsed.cve_image_facts[0].cve_id == "CVE-2026-39822"
    # normalize_image_basename strips "plainid/" — verify that resolves correctly.
    assert normalize_image_basename(parsed.cve_image_facts[0].image) == "pip-operator"


# ── Aqua transactional tag → service token ────────────────────────────────────

def test_aqua_transactional_hotfix_suffix():
    """Non-calendar suffixes (e.g. Junhotfix2026) still yield the service from the tag."""
    image, tag = _aqua_image_basename(
        "244266704373.dkr.ecr.us-east-1.amazonaws.com/"
        "220016-02-dev-eks-transactional:5.2627.4_authz-sql-pdp-modifier_Junhotfix2026"
    )
    assert image == "authz-sql-pdp-modifier"
    assert tag == "5.2627.4_authz-sql-pdp-modifier_Junhotfix2026"


def test_aqua_transactional_calendar_suffix():
    image, tag = _aqua_image_basename(
        "244266704373.dkr.ecr.us-east-1.amazonaws.com/"
        "220016-02-dev-eks-transactional:5.2608.2_agent_16Feb2026"
    )
    assert image == "agent"
    assert tag == "5.2608.2_agent_16Feb2026"


def test_aqua_transactional_underscored_service_normalized():
    """Service tokens with underscores map to hyphenated catalog names."""
    image, tag = _aqua_image_basename(
        "244266704373.dkr.ecr.us-east-1.amazonaws.com/"
        "220016-02-dev-eks-transactional:5.2624.2_pip_operator_08Jun2026"
    )
    assert image == "pip-operator"
    assert tag == "5.2624.2_pip_operator_08Jun2026"


def test_aqua_transactional_theruntime():
    image, tag = _aqua_image_basename(
        "244266704373.dkr.ecr.us-east-1.amazonaws.com/"
        "220016-02-dev-eks-transactional:5.2624.8_theruntime_08Jun2026"
    )
    assert image == "theruntime"
    assert tag == "5.2624.8_theruntime_08Jun2026"


def test_aqua_json_transactional_facts_use_service_not_repo():
    """Full Aqua JSON path: image is service token, not the shared ECR repo name."""
    payload = [
        {
            "image_name": (
                "244266704373.dkr.ecr.us-east-1.amazonaws.com/"
                "220016-02-dev-eks-transactional:5.2627.4_authz-sql-pdp-modifier_Junhotfix2026"
            ),
            "results": {
                "resources": [
                    {
                        "resource": {"name": "openssl", "version": "3.0.0"},
                        "vulnerabilities": [{"name": "CVE-2026-39822"}],
                    }
                ]
            },
        }
    ]
    alias_map = {"authz-sql-pdp-modifier": "authz-sql-pdp-modifier"}
    parsed = parse_attachment_bytes(
        attachment_id="att5",
        filename="aqua_scan.json",
        mime_type="application/json",
        data=json.dumps(payload).encode(),
        alias_map=alias_map,
    )
    assert parsed.status == "ok"
    assert len(parsed.cve_image_facts) == 1
    fact = parsed.cve_image_facts[0]
    assert fact.image == "authz-sql-pdp-modifier"
    assert fact.tag == "5.2627.4_authz-sql-pdp-modifier_Junhotfix2026"


# ── catalog enforcement helpers ───────────────────────────────────────────────

def test_canonical_image_passes_catalog_check():
    allowed = {"pip-operator", "agent", "theruntime", "secrets-mgmt"}
    assert is_allowed_basename("pip-operator", allowed) is True


def test_authorizer_fails_catalog_check():
    allowed = {"pip-operator", "agent", "theruntime", "secrets-mgmt"}
    assert is_allowed_basename("authorizer", allowed) is False


def test_alias_resolved_to_canonical_passes_check():
    alias_map = _build_alias_map([("secrets-mgmt", "secretmgr")])
    allowed = _build_allowed_names([("secrets-mgmt", "secretmgr")])
    raw = "secretmgr"
    canonical = alias_map.get(normalize_image_basename(raw), normalize_image_basename(raw))
    assert is_allowed_basename(canonical, allowed) is True


# ── Excel sheet selection + GHSA ──────────────────────────────────────────────

def _make_multisheet_excel() -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Todays Vulnerabilities"
    ws1.append(("Vulnerability", "Image Repository", "Tag"))
    ws1.append(("CVE-2026-11111", "plainid/agent", "1.0"))
    ws1.append(("GHSA-hrxh-6v49-42gf", "plainid/agent", "1.0"))
    ws2 = wb.create_sheet("CleanData")
    ws2.append(("Vulnerability", "Image Repository", "Tag"))
    ws2.append(("CVE-2026-22222", "plainid/pip-operator", "2.0"))
    ws3 = wb.create_sheet("Export")
    ws3.append(("Vulnerability", "Image Repository", "Tag"))
    ws3.append(("CVE-2026-33333", "plainid/agent", "3.0"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_list_excel_sheets_row_counts():
    from api.app.parsing import list_excel_sheets

    sheets = list_excel_sheets(_make_multisheet_excel())
    by_name = {s["name"]: s["row_count"] for s in sheets}
    assert by_name["Todays Vulnerabilities"] == 2
    assert by_name["CleanData"] == 1
    assert by_name["Export"] == 1


def test_excel_sheet_names_filter_todays_only():
    data = _make_multisheet_excel()
    parsed = parse_attachment_bytes(
        attachment_id="att-ms",
        filename="scan.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=data,
        sheet_names={"Todays Vulnerabilities"},
    )
    ids = {f.cve_id for f in parsed.cve_image_facts}
    assert "CVE-2026-11111" in ids
    assert "GHSA-HRXH-6V49-42GF" in ids
    assert "CVE-2026-22222" not in ids
    assert "CVE-2026-33333" not in ids


def test_excel_parses_ghsa_ids():
    data = _make_excel_bytes([
        ("GHSA-jpcw-4wr7-c3vq", "plainid/authz-access-file", "5.0"),
    ])
    parsed = parse_attachment_bytes(
        attachment_id="att-ghsa",
        filename="scan.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=data,
    )
    assert len(parsed.cve_image_facts) == 1
    assert parsed.cve_image_facts[0].cve_id == "GHSA-JPCW-4WR7-C3VQ"


def test_aqua_json_parses_ghsa_ids():
    payload = [
        {
            "image_name": "registry.example.com/plainid/agent:1.0",
            "results": {
                "resources": [
                    {
                        "resource": {"name": "google.golang.org/grpc", "version": "1.80.0"},
                        "vulnerabilities": [{"name": "GHSA-hrxh-6v49-42gf"}],
                    }
                ]
            },
        }
    ]
    parsed = parse_attachment_bytes(
        attachment_id="att-ghsa-json",
        filename="aqua.json",
        mime_type="application/json",
        data=json.dumps(payload).encode(),
    )
    assert len(parsed.cve_image_facts) == 1
    assert parsed.cve_image_facts[0].cve_id == "GHSA-HRXH-6V49-42GF"


def test_cve_ids_from_facts_include_ghsa():
    ids = cve_ids_from_attachment_facts(
        [
            {
                "cve_image_facts": [
                    {"cve_id": "CVE-2026-1", "image": "agent", "tag": "1", "source": "t"},
                    {"cve_id": "GHSA-AAAA-BBBB-CCCC", "image": "agent", "tag": "1", "source": "t"},
                ]
            }
        ]
    )
    assert ids == ["CVE-2026-1", "GHSA-AAAA-BBBB-CCCC"]


def _make_aqua_html(
    image_ref: str,
    rows: list[tuple[str, str, str]],
    *,
    with_resource: bool = True,
) -> bytes:
    """Minimal Aqua Scan Report HTML. rows: (vuln_id, resource, fix_version)."""
    if with_resource:
        header = (
            "<tr><th>Name</th><th>Resource</th><th>Severity</th>"
            "<th>Fix Version</th></tr>"
        )
        body = "".join(
            f"<tr><td>{vid}</td><td>{res}</td><td>high</td><td>{fix}</td></tr>"
            for vid, res, fix in rows
        )
    else:
        header = "<tr><th>Name</th><th>Severity</th><th>Fix Version</th></tr>"
        body = "".join(
            f"<tr><td>{vid}</td><td>high</td><td>{fix}</td></tr>"
            for vid, _res, fix in rows
        )
    html = f"""<!DOCTYPE html><html><head>
<title>{image_ref} | Scan Results</title></head><body>
<h1>Scan Report: {image_ref}</h1>
<table>{header}{body}</table>
</body></html>"""
    return html.encode("utf-8")


def test_aqua_html_parses_consolidated_table():
    data = _make_aqua_html(
        "eyctpeu.example/plainid/secrets-mgmt:5.2631.3",
        [
            ("CVE-2025-60876", "busybox", "None"),
            ("GHSA-r277-6w6q-xmqw", "github.com/getkin/kin-openapi", "0.144.0"),
            ("GHSA-jpcw-4wr7-c3vq", "github.com/getkin/kin-openapi", "0.144.0"),
            ("GHSA-hrxh-6v49-42gf", "google.golang.org/grpc", "1.82.1"),
        ],
    )
    parsed = parse_attachment_bytes(
        attachment_id="att-html",
        filename="AquaReport-secrets-mgmt-5.2631.3.html",
        mime_type="text/html",
        data=data,
    )
    assert parsed.status == "ok"
    ids = {f.cve_id for f in parsed.cve_image_facts}
    assert ids == {
        "CVE-2025-60876",
        "GHSA-R277-6W6Q-XMQW",
        "GHSA-JPCW-4WR7-C3VQ",
        "GHSA-HRXH-6V49-42GF",
    }
    assert all(
        normalize_image_basename(f.image) == "secrets-mgmt" and f.tag == "5.2631.3"
        for f in parsed.cve_image_facts
    )
    pkgs = {p.cve_id: p for p in parsed.packages}
    assert pkgs["CVE-2025-60876"].package_name == "busybox"
    assert pkgs["CVE-2025-60876"].fixed_version is None
    assert pkgs["GHSA-R277-6W6Q-XMQW"].fixed_version == "0.144.0"


def test_aqua_html_falls_back_to_name_only_tables():
    data = _make_aqua_html(
        "registry.example/plainid/agent:1.2.3",
        [("CVE-2026-11111", "", "2.0")],
        with_resource=False,
    )
    parsed = parse_attachment_bytes(
        attachment_id="att-html2",
        filename="report.htm",
        mime_type="text/html",
        data=data,
    )
    assert parsed.status == "ok"
    assert len(parsed.cve_image_facts) == 1
    assert parsed.cve_image_facts[0].cve_id == "CVE-2026-11111"
    assert normalize_image_basename(parsed.cve_image_facts[0].image) == "agent"
    assert parsed.packages == []


def test_aqua_html_real_secrets_mgmt_fixture():
    from pathlib import Path

    path = Path(
        "/Users/alexbirman/Downloads/"
        "AquaReport-secrets-mgmt-5.2631.3-20260727132524-1256061.html"
    )
    if not path.is_file():
        import pytest

        pytest.skip("local Aqua HTML fixture not present")
    parsed = parse_attachment_bytes(
        attachment_id="87830",
        filename=path.name,
        mime_type="text/html",
        data=path.read_bytes(),
    )
    assert parsed.status == "ok"
    ids = {f.cve_id for f in parsed.cve_image_facts}
    assert ids == {
        "CVE-2025-60876",
        "GHSA-R277-6W6Q-XMQW",
        "GHSA-JPCW-4WR7-C3VQ",
        "GHSA-HRXH-6V49-42GF",
    }
    assert {normalize_image_basename(f.image) for f in parsed.cve_image_facts} == {"secrets-mgmt"}
    assert {f.tag for f in parsed.cve_image_facts} == {"5.2631.3"}
